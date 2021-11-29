import os
import re
from openpyxl import load_workbook
def main():
    filename = fileBl()
    shuju = excelcz(filename)
def fileBl():
    #获取文件数量，返回文件名列表
    path = "F:/jftest"   #指定文件路径
    for root, dirs, files in os.walk(path, True):#进行目录遍历
        print ('root: %s' % root)#输出文件所属目录
        print ('dirs: %s' % dirs)#输出子文件夹
        print ('files:{}'.format(files))#输出文件名
        return(files)
def excelcz(list1):
    workbook1 = load_workbook('F:/jftest.xlsx')
    sheet1 = workbook1['Sheet1']
    filecount = len(list1)
    for i in range(filecount):
        filename1=list1[i]
        workbook2 = load_workbook('F:/jftest/'+list1[i])
        sheet2 =  workbook2['团队客户 - 第1表']
        row1 = sheet2.max_row
        column1 = sheet2.max_column
        for j in range(1,row1+1):
            list2=[]
            for k in range(1,column1+1):
                if j==1 and k==1:
                    list2.append(filename1)
                    sheet1.append(list2)
                    list2=[]
                value=sheet2.cell(row=j,column=k).value
                list2.append(value)
            sheet1.append(list2)
    workbook1.save('F:/jftest.xlsx')
if __name__ == '__main__':
    main()