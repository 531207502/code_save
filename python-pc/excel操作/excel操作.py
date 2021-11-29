from openpyxl import load_workbook
#此处获得的workbook是一个操作这个excel的对象
workbook=load_workbook('F:/test/test1.xlsx')
print(workbook.sheetnames)
#此处sheet1是根据对应的分页标签名获得的标签
sheet1=workbook['cisco']
print(sheet1)
print(type(sheet1))
print("下面的是行和列的参数")
print(sheet1.max_row)
print(sheet1.max_column)
#两种方式获取单元格数据，坐标，或者行列数,返回的是类，要通过value来获取值
cell1=sheet1['A1']
print(cell1)
print(type(cell1))
print(cell1.value)
cell2=sheet1.cell(row=2,column=2)
print(cell2)
print(type(cell2))
print(cell2.value)
#获取多行值,可以按照坐标范围获取，如下，获取的时候他还是按行把数据放在一个列表中，也可以按照列或者行指定获取
cell3=sheet1['A1:C4']
print(type(cell3))
print(cell3)
for i in cell3:
    num=len(i)
    for j in range(num):
        print(i[j].value)
#按列获取值
cell4=sheet1['A:C']
print(cell4)
c=1
for k in cell4:
    m=len(k)
    for n in range(m):
        print(k[n].value)
        k[n].value=c
        c+=1
#追加数据
list1=["abc","bcd","cde","def"]
sheet1.append(list1)
#插入行，注意是在idx行上面插入,插入多行
sheet1.insert_rows(idx=2)
sheet1.insert_rows(idx=4,amount=4)
#插入列，也和插入行概念一样
sheet1.insert_cols(idx=3)
#删除行和列，删除的时候包含idx这一行或者列，书写格式一样，只是变成delete_cols和delete_rows
#保存excel
workbook.save('F:/test/test1.xlsx')