import re
import openpyxl
def textcz(filename,textcount):
    for i in range(len(textcount)):
        result=re.search('Processor Pool Total:(.*) Used:(.*) Free:(.*)',textcount[i])
        result3=re.search('System Memory : (.*)K total, (.*)K used, 1(.*)K free',textcount[i])
        result4 = re.search('Memory usage:   (.*)K total,   (.*)K used,   (.*)K free', textcount[i])
        if result:
            list1=[filename,result.group(1),result.group(2),result.group(3)]
            num=i+1
            result1=re.search('(lsmpi_io|I/O) Pool Total:(.*) Used:(.*) Free:(.*)',textcount[num])
            result2 = re.search('Driver te Pool Total:(.*) Used:(.*) Free:(.*)', textcount[num + 1])
            if result2:
                list2 = [filename,result1.group(2),result1.group(3),result1.group(4)]
                list3 = [filename,result2.group(1), result2.group(2), result2.group(3)]
                writeexcel2(list1, list2,list3)
            else:
                list2 = [filename,result1.group(2),result1.group(3),result1.group(4)]
                writeexcel1(list1,list2)
            break
        elif result3:
            list1 = [filename, result3.group(1), result3.group(2), result3.group(3)]
            writeexcel3(list1)
            break
        elif result4:
            list1 = [filename, result4.group(1), result4.group(2), result4.group(3)]
            writeexcel3(list1)
            break
def writeexcel1(li1,li2):
    workbook=openpyxl.load_workbook('D:/xunjian/cisco.xlsx')
    memory1=workbook['memory1']
    memory1.append(li1)
    memory1.append(li2)
    workbook.save('D:/xunjian/cisco.xlsx')
def writeexcel2(li1,li2,li3):
    workbook=openpyxl.load_workbook('D:/xunjian/cisco.xlsx')
    memory1=workbook['memory1']
    memory1.append(li1)
    memory1.append(li2)
    memory1.append(li3)
    workbook.save('D:/xunjian/cisco.xlsx')
def writeexcel3(li1):
    workbook=openpyxl.load_workbook('D:/xunjian/cisco.xlsx')
    memory1=workbook['memory1']
    memory1.append(li1)
    workbook.save('D:/xunjian/cisco.xlsx')