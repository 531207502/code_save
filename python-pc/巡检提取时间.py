from time import sleep
import openpyxl
import os
import re
def tqtime(file_object,sh,nr):
    global row1
    x = -1
    n = 0
    while lines - 1 != x:  # 逐行匹配
        x += 1
        a = re.search("uptime is *", nr[x])
        if a == None:
            print(nr[x])
            #sleep(1)
            continue
        else:
            sh.cell(row=row1, column=2, value=nr[x])
            break
def filename(file_object,sh,nr):
    global row1
    x = -1
    n = 0
    while lines - 1 != x:  # 逐行匹配
        x += 1
        a = re.findall("^<.*>display", nr[x])
        b = re.findall(".*#show version", nr[x])
        if a == [] and b==[]:
            continue
        elif a!=[]:
            #print(a)
            sh.cell(row=row1, column=1, value=a[0])
            break
        elif b!=[]:
            sh.cell(row=row1, column=1, value=b[0])
            break
path = "F:/cisco"   #指定文件路径
for root, dirs, files in os.walk(path, True):#进行目录遍历
    print ('root: %s' % root)#输出文件所属目录
    print ('dirs: %s' % dirs)#输出子文件夹
    #print ('files: %s' % files)#输出文件名
    print ('')
d = 0
wenjianshu = len(files)#统计总共文件数量
while wenjianshu != d:
    files[d] = "F:/cisco/" + files[d]#生成文件完整路径
    print(files[d])#打印出路径，可以不要这句
    d = d + 1#根据文件数量来进行循环
e = 0
wb = openpyxl.Workbook()
sh = wb["Sheet"]
row1 = 1
while e != wenjianshu:
    with open(files[e], encoding="utf-8")as file_object:#一次打开一个文件，依次打开文件
        nr = file_object.readlines()  # 读取文件内容
        lines = len(nr)  # 统计文件行数
        filename(file_object, sh,nr)
        tqtime(file_object,sh,nr)
        row1+=1
        e+=1
wb.save(r"F:/cisco/test111.xlsx")
